CL_MODEL = "Curriculum Yes ob easy 0_7"
CL_KIND = f"{CL_MODEL}_test_0"
TEST_COUNT = 10
INPUT_SENSOR = 8
INPUT_SIZE = 10
NORMALIZATION_SENSOR = 100
MAX_SPEED = 1.57
COLLISION_R = 6

from controller import Supervisor
import matplotlib.pyplot as plt
import os
import datetime
import tensorflow as tf
from tensorflow import keras
import numpy as np
import math

DAY_NUM = str(datetime.date.today())

# 1. 기초 세팅
robot = Supervisor()
# 1-1. model_load (model name is 20230315_model)
Dqn_model = tf.keras.models.load_model(CL_MODEL)
# 1-1. model detail
# new_model.summary()
# 1-2. 현재 world timestep
timestep = int(robot.getBasicTimeStep())
# 1-3. 로봇 모터 정보
left_motor = robot.getDevice('left wheel motor')
right_motor = robot.getDevice('right wheel motor')


# 1-3. 로봇 모터 다음 명령 있을 때 까지 전 모터 상태 유지
left_motor.setPosition(float('inf'))
right_motor.setPosition(float('inf'))
# 1-4. 로봇 노드 정보

ep_node = robot.getFromDef('e-puck')
# 1-3. 로봇 필드 정보
translation_field = ep_node.getField('translation')
rotation_field = ep_node.getField('rotation')
# 1-3-1. 장애물 필드 정보

#1-6. 로봇 적외선 센서 정보
ps = []
psNames = ['ps0', 'ps1', 'ps2', 'ps3','ps4', 'ps5', 'ps6', 'ps7']
for i in range(8):
    ps.append(robot.getDevice(psNames[i]))
    ps[i].enable(timestep)
# 1-7. 초기화
storage = []
state = np.zeros((30))
count_state = 0
ep = []
result_collision_done = []
finish = 0
# 1-7. goal 초기화

goal = [0,0,0]
print("goal :",goal)

x_min, x_max = -0.9, 0.9
y_min, y_max = -0.9, 0.9

x = x_min
y = y_min
xy_count = 0
time_count = 0
done_count = 0
collision_count = 0
localminima_count = 0

result_done = []
result_collision = []
result_minima = []



def draw_trajectories(trajectories):
    colors = plt.cm.rainbow(np.linspace(0, 1, len(trajectories)))
    for i, trajectory in enumerate(trajectories):
        plt.plot(trajectory[:, 0], trajectory[:, 1], color=colors[i], label=f'Trajectory {i+1}')
    
    # 4. 결과 저장
    createDirectory(f"data")
    createDirectory(f"data/{CL_MODEL}")
    createDirectory(f"data/{CL_MODEL}/{CL_KIND}")
    plt.savefig(f'data/{CL_MODEL}/{CL_KIND}_Tragectory.png')
    plt.cla()
    
# 2. define function
# 2-1. e-puck  perfect_angle
def point_slope(target):
    ep = translation_field.value
    slope = (target[1] - ep[1]) / (target[0] - ep[0])
    if slope < 0:
        result = np.degrees(math.atan(slope))+180
        if target[0] < ep[0] and target[1] > ep[1]:
            return result
        elif target[0] > ep[0] and target[1] < ep[1]:
            return result + 180
        elif target[0] > ep[0] and target[1] > ep[1]:
            return result
        elif target[0] < ep[0] and target[1] < ep[1]:
            return result + 180
        else:
            return result

    else:
        result = np.degrees(math.atan(slope))
        if target[0] < ep[0] and target[1] > ep[1]:
            return result
        elif target[0] > ep[0] and target[1] < ep[1]:
            return (result + 180)
        elif target[0] > ep[0] and target[1]> ep[1]:
            return result 
        elif target[0] < ep[0] and target[1] < ep[1]:
            return (result + 180)
        else:
            return result
# 2-2. coordinate transformation
def rotated_point(orientation):
    angle = np.radians(-90)
    rot_matrix = np.array([[np.cos(angle), -np.sin(angle), 0],
                          [np.sin(angle), np.cos(angle), 0],
                          [0, 0, 1]])
    point = np.array(orientation[:3])
    rotated_point = np.dot(rot_matrix,point)
    heading = math.atan2(rotated_point[0], rotated_point[1])
    heading_degrees = np.degrees(heading) + 180
    return heading_degrees
# 2-3. select action 
def Action(action):
    # Go straight
    if action == 0:
        left_motor.setVelocity(MAX_SPEED)
        right_motor.setVelocity(MAX_SPEED)
    # Trun Right
    elif action == 1:
        left_motor.setVelocity(MAX_SPEED)
        right_motor.setVelocity(MAX_SPEED/3)
    # Trun Left
    elif action == 2:
        left_motor.setVelocity(MAX_SPEED/3)
        right_motor.setVelocity(MAX_SPEED)
    # Trun Right
    elif action == 3:
        left_motor.setVelocity(MAX_SPEED)
        right_motor.setVelocity(-MAX_SPEED)
    # Trun Left
    elif action == 4:
        left_motor.setVelocity(-MAX_SPEED)
        right_motor.setVelocity(MAX_SPEED)
# 0.3925
# 2-4. state get
def environment():
    # 2-1-1. location get
    ep = translation_field.value
    # 2-1-2. heading get
    orientation = ep_node.getOrientation()
    heading = rotated_point(orientation)
    # 2-1-3. perfect_angle get    
    perfect_angle = point_slope(goal)                           
    # 2-1-4. theta get
    theta = heading - perfect_angle
    if abs(theta) > 180:
        if theta < 0:
            theta = theta + 360
        elif theta > 0:
            theta = theta - 360
    # 2-1-5. radius get ep 
    goal_radius = math.sqrt(pow(goal[0] - ep[0],2) + pow(goal[1] - ep[1],2))
    # 2-1-6. radius get ob
    storage.append(goal_radius)
    storage.append(math.radians(theta))
    # 2-1-6-1. sensor value
    for i in range(INPUT_SENSOR):
        storage.append(ps[i].value/NORMALIZATION_SENSOR)

# 2-6. policy
def policy(state):                                                                                # state는 1차 리스트고 찐 policy 임.
    state_input = tf.convert_to_tensor(state[None, :], dtype=tf.float32)                                # tensor 형태로 변환
    action_q = Dqn_model(state_input)                      
    action = np.argmax(action_q.numpy()[0], axis=0)                                                     # action은 0,1 이런식으로 나옴.               
    return action
    
# 2.7 collision check
def collision_check():
    global time_count 
    global collision_count
    global result_collision_done
    global trajectory
    for j in range(3):
        for i in range(2,INPUT_SENSOR + 2):
            if (action == 1
            or action == 2
            or i == 5
            or i == 6):
                continue
            if COLLISION_R < state[j * INPUT_SIZE + i] :
                collision_count += 1
                time_count = 0
                setting()
                result_collision_done.append(-1)
                result_collision.append(1)
                trajectory = []
                return

def setting():
    global x
    global y
    global xy_count
    print("Current : ",len(result_collision_done))
    print("xy_count :",xy_count)
    print("x :",x)
    print("y :",y)
    result_done.append(0)
    result_collision.append(0)
    result_minima.append(0)
    if x >= 0.89 and y <= -0.89 and xy_count == 0:
        xy_count = 1
    elif x >= 0.89 and y >= 0.89 and xy_count == 1:
        xy_count = 2
    elif x <= -0.89 and y >= 0.89 and xy_count == 2:
        xy_count = 3
    elif x < -0.89 and y < -0.89 and xy_count == 3:
        xy_count = 1
    
    if xy_count == 0:
        x += (1.8 / TEST_COUNT)
    elif xy_count == 1:
        y += (1.8 / TEST_COUNT)
    elif xy_count == 2:
        x -= (1.8 / TEST_COUNT)
    elif xy_count == 3:
        y -= (1.8 / TEST_COUNT)
    r1 = np.random.uniform(-3,3)
    translation_field.setSFVec3f([translation_field.value[0],translation_field.value[1],5])
    translation_field.setSFVec3f([x,y,5])
    translation_field.setSFVec3f([x,y,0])
    rotation_field.setSFRotation([0,0,-1,r1])
    robot.simulationResetPhysics()
    return

def createDirectory(directory):
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
    except OSError:
        print("Error: Failed to create the directory.")


# 3. collision avoidance
setting()
trajectory = []
trajectories = []

while robot.step(timestep) != -1:
    count_state += 1  
    environment()
    if count_state == 3:
        count_state = 0  
        # 현 state 가져오기.    
        state = np.array(storage)
        # storage 초기화 하고
        storage = []
        # 3-1-2. action 하고 다음 state로 넘어감.
        action = policy(state)
        Action(action)
        # collision check
        collision_check()
        trajectory.append(list(translation_field.value[0:2]))
        for i in range(3):
            if state[i * INPUT_SIZE] < 0.1:
                trajectories.append(np.array(trajectory))
                trajectory = []
                result_done.append(1)
                done_count += 1
                time_count = 0
                result_collision_done.append(1)
                setting()
                break
        finish = finish + 1
        time_count += 1
        if time_count == 3000:
            trajectory = []
            time_count = 0
            localminima_count += 1
            result_collision_done.append(0)
            result_minima.append(1)
            setting()
        
    if len(result_collision_done) >= TEST_COUNT * 4:
        draw_trajectories(trajectories)
        break
        
x_data = list(range(len(result_collision_done)))
loss_min = -1.1
loss_max = 1.1
plt.ylim([loss_min, loss_max])
plt.xlabel('Episode')
plt.ylabel('Done')
plt.plot(x_data,result_collision_done,c='red',label = "Done_Collision_Minima")
plt.savefig(f'data/{CL_MODEL}/{CL_KIND}/{CL_KIND}_Done_Collision_Minima.png')
plt.cla()

plt.ylim([-0.1, loss_max])
plt.xlabel('Episode')
plt.ylabel('Done')
plt.plot(list(range(len(result_done))),result_done,c='green',label = "Done")
plt.savefig(f'data/{CL_MODEL}/{CL_KIND}/{CL_KIND}_Done_result.png')
plt.cla()

plt.ylim([-0.1, loss_max])
plt.xlabel('Episode')
plt.ylabel('Collision')
plt.plot(list(range(len(result_collision))),result_collision,c='blue',label = "Collision")
plt.savefig(f'data/{CL_MODEL}/{CL_KIND}/{CL_KIND}_Collision_result.png')
plt.cla()

plt.ylim([-0.1, loss_max])
plt.xlabel('Episode')
plt.ylabel('Minima')
plt.plot(list(range(len(result_minima))),result_minima,c='black',label = "Local minima")
plt.savefig(f'data/{CL_MODEL}/{CL_KIND}/{CL_KIND}_Minima_result.png')
plt.cla()


from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws['A1'] = '에피소드 횟수'
ws['A2'] = len(result_collision_done)
ws['B1'] = '도착 횟수'
ws['B2'] = done_count
ws['C1'] = '충돌 횟수'
ws['C2'] = collision_count
ws['D1'] = 'Local minima 횟수'
ws['D2'] = localminima_count

ws['B4'] = '도착 확률'
ws['B5'] = done_count/len(result_collision_done)
ws['C4'] = '충돌 확률'
ws['C5'] = collision_count/len(result_collision_done)
ws['D4'] = 'Minima 확률'
ws['D5'] = localminima_count/len(result_collision_done)

wb.save(f'data/{CL_MODEL}/{CL_KIND}/{CL_KIND}_Test_Result.xlsx')    
    
